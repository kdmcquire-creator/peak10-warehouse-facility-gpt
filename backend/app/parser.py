from io import BytesIO
from collections import defaultdict
from datetime import datetime
from uuid import uuid4
from openpyxl import load_workbook
from .models import ParsedWorkbook, WorkbookTabStatus, BaseMonthlyRow, HedgePosition

REQUIRED_TABS = ["CF", "Strip Pricing", "1-month Term SOFR", "GRC Hedges", "Brown Pony Hedges"]

def _to_date_key(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value) if value is not None else ""

def _header_map(ws, header_row=1):
    row = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    return {str(v).strip(): idx for idx, v in enumerate(row, start=1) if v is not None}

def parse_cf(ws):
    headers = _header_map(ws, 1)
    buckets = defaultdict(lambda: {
        "gr_oil": 0.0, "gr_gas": 0.0, "n_oil": 0.0, "n_gas": 0.0, "n_ngl": 0.0,
        "n_tot_rev": 0.0, "opinc": 0.0, "n_capex": 0.0
    })
    for row in ws.iter_rows(min_row=2, values_only=True):
        rsv = row[headers["RSV_CAT"] - 1] if "RSV_CAT" in headers else None
        scenario = row[headers["SCENARIO"] - 1] if "SCENARIO" in headers else None
        if rsv != "1PDP" or scenario != "PK10":
            continue
        outdate = _to_date_key(row[headers["OUTDATE"] - 1])
        b = buckets[outdate]
        mapping = {
            "GR OIL": "gr_oil", "GR GAS": "gr_gas", "N OIL": "n_oil", "N GAS": "n_gas",
            "N NGL": "n_ngl", "N TOT REV": "n_tot_rev", "OPINC": "opinc", "N CAPEX": "n_capex"
        }
        for src, dst in mapping.items():
            if src in headers:
                b[dst] += float(row[headers[src] - 1] or 0.0)

    results = []
    for outdate in sorted(buckets.keys()):
        b = buckets[outdate]
        fcf = b["opinc"] - b["n_capex"]
        results.append(BaseMonthlyRow(
            outdate=outdate,
            gr_oil=b["gr_oil"],
            gr_gas=b["gr_gas"],
            n_oil=b["n_oil"],
            n_gas=b["n_gas"],
            n_ngl=b["n_ngl"],
            n_tot_rev=b["n_tot_rev"],
            opinc=b["opinc"],
            n_capex=b["n_capex"],
            fcf_unhedged=fcf,
            fcf_hedged=fcf,
        ))
    return results

def parse_strip(ws):
    headers = _header_map(ws, 4)
    data = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        eom = row[headers["EOMONTH"] - 1] if "EOMONTH" in headers else None
        if eom is None:
            continue
        key = _to_date_key(eom)
        data[key] = {
            "wti": float(row[headers["NYMEX WTI CMA"] - 1] or 0.0),
            "hh": float(row[headers["NYMEX Henry Hub (LD)"] - 1] or 0.0),
            "waha": float(row[headers["Waha Basis"] - 1] or 0.0),
        }
    return data

def parse_sofr(ws):
    headers = _header_map(ws, 7)
    data = {}
    date_header = "EoMonth" if "EoMonth" in headers else ("Date" if "Date" in headers else None)
    if date_header is None or "SOFR" not in headers:
        return data
    for row in ws.iter_rows(min_row=8, values_only=True):
        date_val = row[headers[date_header] - 1]
        sofr = row[headers["SOFR"] - 1]
        if date_val is None or sofr is None:
            continue
        data[_to_date_key(date_val)] = float(sofr or 0.0)
    return data

def parse_hedges(ws):
    headers = _header_map(ws, 1)
    positions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        end_date = row[headers["Contract End Date"] - 1] if "Contract End Date" in headers else None
        if end_date is None:
            continue
        positions.append(HedgePosition(
            entity=str(row[headers["Entity"] - 1]) if "Entity" in headers and row[headers["Entity"] - 1] is not None else None,
            underlying=str(row[headers["Underlying"] - 1]) if "Underlying" in headers and row[headers["Underlying"] - 1] is not None else None,
            trade_type=str(row[headers["Trade Type"] - 1]) if "Trade Type" in headers and row[headers["Trade Type"] - 1] is not None else None,
            direction=str(row[headers["Direction"] - 1]) if "Direction" in headers and row[headers["Direction"] - 1] is not None else None,
            flavor=str(row[headers["Flavor"] - 1]) if "Flavor" in headers and row[headers["Flavor"] - 1] is not None else None,
            strike=float(row[headers["Price/Strike"] - 1] or 0.0) if "Price/Strike" in headers else None,
            quantity=float(row[headers["Quantity"] - 1] or 0.0) if "Quantity" in headers else None,
            contract_end_date=_to_date_key(end_date),
        ))
    return positions

def calc_hedge_payoff(position, market):
    if position.underlying == "NYMEX WTI CMA":
        px = market.get("wti", 0.0)
    elif position.underlying == "NYMEX Henry Hub (LD)":
        px = market.get("hh", 0.0)
    elif position.underlying == "Waha Basis":
        px = market.get("waha", 0.0)
    else:
        return 0.0
    strike = float(position.strike or 0.0)
    qty = abs(float(position.quantity or 0.0))
    if position.trade_type == "Swaps":
        return (strike - px) * qty
    if position.trade_type == "Options" and position.flavor == "Put" and position.direction == "Buy":
        return max(strike - px, 0.0) * qty
    if position.trade_type == "Options" and position.flavor == "Call" and position.direction == "Sell":
        return -max(px - strike, 0.0) * qty
    return 0.0

def parse_workbook(content: bytes, file_name: str) -> ParsedWorkbook:
    wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    tabs = [WorkbookTabStatus(name=t, found=t in wb.sheetnames) for t in REQUIRED_TABS]
    notes = [
        "Current-base replication uses CF rows where RSV_CAT == 1PDP and SCENARIO == PK10.",
        "Existing hedge payoffs are recomputed monthly from parsed hedge positions and strip prices.",
        "Strip and 1-month Term SOFR are joined by month-end date."
    ]

    cf_rows = parse_cf(wb["CF"]) if "CF" in wb.sheetnames else []
    strip = parse_strip(wb["Strip Pricing"]) if "Strip Pricing" in wb.sheetnames else {}
    sofr = parse_sofr(wb["1-month Term SOFR"]) if "1-month Term SOFR" in wb.sheetnames else {}
    hedges = []
    if "GRC Hedges" in wb.sheetnames:
        hedges.extend(parse_hedges(wb["GRC Hedges"]))
    if "Brown Pony Hedges" in wb.sheetnames:
        hedges.extend(parse_hedges(wb["Brown Pony Hedges"]))

    hedge_by_month = defaultdict(list)
    for h in hedges:
        hedge_by_month[h.contract_end_date].append(h)

    enriched = []
    for row in cf_rows:
        market = strip.get(row.outdate, {})
        row.wti = market.get("wti")
        row.hh = market.get("hh")
        row.waha = market.get("waha")
        row.sofr = sofr.get(row.outdate)
        payoff = sum(calc_hedge_payoff(h, market) for h in hedge_by_month.get(row.outdate, []))
        row.hedge_payoff = payoff
        row.fcf_hedged = row.fcf_unhedged + payoff
        enriched.append(row)

    model_id = uuid4().hex
    return ParsedWorkbook(
        model_id=model_id,
        file_name=file_name,
        tabs=tabs,
        notes=notes,
        month_count=len(enriched),
        first_month=enriched[0].outdate if enriched else None,
        last_month=enriched[-1].outdate if enriched else None,
        base_monthly=enriched,
        hedges=hedges,
    )
